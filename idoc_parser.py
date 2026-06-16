#!/usr/bin/env python3
"""
SAP IDoc flat-file → Excel documentation generator.
Usage: python idoc_parser.py <input.txt> [output.xlsx]
"""
import sys
import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Field definitions ──────────────────────────────────────────────────────
SEGMENT_FIELDS = {

    # ── Short/specific segments (must be defined before generic E2EDK*/E2EDP*) ──
    'E2EDK02': [
        ('QUALF',3,'Qualifier'),('BELNR',35,'Document number'),('POSNR',6,'Item number'),
        ('DATUM',8,'Date'),('UZEIT',6,'Time'),
    ],
    'E2EDK03': [
        ('IDDAT',3,'Date qualifier'),('DATUM',8,'Date'),('UZEIT',6,'Time'),
    ],
    'E2EDK14': [
        ('QUALF',3,'Qualifier'),('ORGID',35,'Organizational unit ID'),
    ],
    'E2EDK17': [
        ('QUALF',3,'Qualifier'),('LKOND',3,'Loading condition'),('LKTEXT',70,'Loading condition description'),
    ],
    'E2EDK18': [
        ('QUALF',3,'Qualifier'),('TAGE',8,'Days'),('PRZNT',8,'Percentage'),('ZTERM_TXT',70,'Payment terms description'),
    ],
    'E2EDK35': [
        ('QUALZ',3,'Qualifier'),('CUSADD',35,'Additional data'),('CUSADD_BEZ',40,'Additional data description'),
    ],
    'E2EDP17': [
        ('QUALF',3,'Qualifier'),('LKOND',3,'Loading condition'),('LKTEXT',70,'Delivery condition description'),('LPRIO',2,'Delivery priority'),
    ],
    'E2EDP19': [
        ('QUALF',3,'Qualifier'),('IDTNR',35,'Object ID (material number)'),('KTEXT',70,'Short text / description'),
        ('MFRPN',42,'Manufacturer part number'),('MFRNR',10,'Manufacturer number'),
        ('IDTNR_EXTERNAL',40,'Object ID external'),('IDTNR_VERSION',10,'Version number for IDTNR'),
        ('IDTNR_GUID',32,'External GUID for IDTNR'),('IDTNR_LONG',40,'Object ID (long)'),
    ],
    'E2EDP20': [
        ('WMENG',15,'Scheduled quantity'),('AMENG',15,'Previous quantity'),
        ('EDATU',8,'Delivery date'),('EZEIT',6,'Delivery time'),
        ('EDATU_OLD',8,'Previous delivery date'),('EZEIT_OLD',6,'Previous delivery time'),
        ('ACTION',3,'Action code'),('HANDOVERDATE',8,'Handover date'),('HANDOVERTIME',6,'Handover time'),
    ],
    'E2EDKT1': [
        ('TDID',4,'Text ID'),('TSSPRAS',3,'Language key'),('TSSPRAS_ISO',2,'2-character SAP language code'),
        ('TDOBJECT',10,'Texts: application object'),('TDOBNAME',70,'Name'),
    ],
    'E2EDKT2': [
        ('TDLINE',70,'Text line'),('TDFORMAT',2,'Tag column'),
    ],
    'E2EDPT1': [
        ('TDID',4,'Text ID'),('TSSPRAS',3,'Language key'),('TSSPRAS_ISO',2,'2-character SAP language code'),
    ],
    'E2EDPT2': [
        ('TDLINE',70,'Text line'),('TDFORMAT',2,'Tag column'),
    ],
    'E2EDS01': [
        ('SUMID',3,'Total qualifier'),('SUMME',18,'Total value'),
        ('SUNIT',3,'Currency unit'),('WAERQ',3,'Currency'),
    ],
    'E2ADDI': [
        ('ADDIMATNR',18,'Sales promotion material number'),
        ('ADDINUMBER',17,'Number of sales promotion items'),
        ('ADDIVKME',3,'Sales unit of measure of material'),
        ('ADDIFM',4,'Attachment method for sales promotion items'),
        ('ADDIFM_TXT',40,'Sales promotion items: description of attachment method'),
        ('ADDIKLART',3,'Class type for display when editing sales promotion items'),
        ('ADDIKLART_TXT',40,'Description of class type'),
        ('ADDICLASS',18,'Class whose elements are assigned to sales promotion items'),
        ('ADDICLASS_TXT',40,'Keywords'),
        ('ADDIIDOC',1,'Reference indicator for separate sales promotion IDoc'),
        ('ADDIMATNR_EXTERNAL',40,'Long material number for field ADDIMATNR'),
        ('ADDIMATNR_VERSION',10,'Version number for field ADDIMATNR'),
        ('ADDIMATNR_GUID',32,'External GUID for field ADDIMATNR'),
        ('ADDIMATNR_LONG',40,'Sales promotion material number'),
    ],
    'E2CUBLB': [
        ('CONTENT',250,'CU: BLOB of a configuration (SCE)'),
    ],
    'E2CUCFG': [
        ('POSEX',6,'External item number'),
        ('CONFIG_ID',6,'External configuration ID (temporary)'),
        ('ROOT_ID',8,'Instance number in configuration'),
        ('SCE',1,'CU: information passed only by SCE for configuration'),
        ('KBNAME',30,'Knowledge base object'),
        ('KBVERSION',30,'Runtime version of an SCE knowledge base'),
        ('COMPLETE',1,'General indicator'),
        ('CONSISTENT',1,'General indicator'),
        ('CFGINFO',250,'CU: BLOB of a configuration (SCE)'),
        ('KBPROFILE',30,'Knowledge base profile'),
        ('KBLANGUAGE',1,'Language of a configuration'),
        ('CBASE_ID',32,'Instance number (persistent)'),
        ('CBASE_ID_TYPE',1,'Type of instance number (persistent)'),
    ],
    'E2CUINS': [
        ('INST_ID',8,'Instance number in configuration'),
        ('OBJ_TYPE',10,'Object type'),
        ('CLASS_TYPE',3,'Class type'),
        ('OBJ_KEY',50,'Object key'),
        ('OBJ_TXT',70,'Language-dependent object description'),
        ('QUANTITY',15,'Instance quantity'),
        ('AUTHOR',1,'Statement was derived'),
        ('QUANTITY_UNIT',3,'Unit of measure'),
        ('COMPLETE',1,'General indicator'),
        ('CONSISTENT',1,'General indicator'),
        ('OBJECT_GUID',32,'GUID for TYPE_OF statement of instance'),
        ('PERSIST_ID',32,'Instance number (persistent)'),
        ('PERSIST_ID_TYPE',1,'Type of instance number (persistent)'),
    ],
    'E2CUPRT': [
        ('PARENT_ID',8,'Character field of length 8'),
        ('INST_ID',8,'Character field of length 8'),
        ('PART_OF_NO',4,'Undefined area, possibly usable for patch levels'),
        ('OBJ_TYPE',10,'Character field of length 10'),
        ('CLASS_TYPE',3,'Field of length 3 bytes'),
        ('OBJ_KEY',50,'Comment'),
        ('AUTHOR',1,'Statement was derived'),
        ('SALES_RELEVANT',1,'Part is sales-relevant'),
        ('PART_OF_GUID',32,'GUID for PART_OF statement of instance'),
    ],
    'E2CUREF': [
        ('POSEX',6,'Character field of length 6'),
        ('CONFIG_ID',6,'Character field of length 6'),
        ('INST_ID',8,'Character field of length 8'),
    ],
    'E2CUVAL': [
        ('INST_ID',8,'Character field of length 8'),
        ('CHARC',40,'Character field 40 characters'),
        ('CHARC_TXT',70,'Character field of length 70'),
        ('VALUE',40,'Character field 40 characters'),
        ('VALUE_TXT',70,'Character field of length 70'),
        ('AUTHOR',1,'Statement was derived'),
        ('VALUE_TO',40,'Value of a characteristic'),
        ('VALCODE',1,'Value type: interval limits - single value'),
        ('VALUE_LONG',70,'Characteristic value length 70'),
        ('VALUE_TO_LONG',70,'Characteristic value length 70'),
    ],
    'E2EDC': [
        ('SGTYP',3,'IDoc service specification segment type'),
        ('ZLTYP',3,'IDoc service specification line type'),
        ('LVALT',3,'IDoc service specification alternatives'),
        ('ALTNO',2,'IDoc alternative number for service specifications'),
        ('ALREF',5,'IDoc assignment number for service specifications'),
        ('ZLART',3,'IDoc service specification line category'),
        ('POSEX',10,'Line number'),
        ('RANG',2,'Hierarchy level of group'),
        ('EXGRP',8,'Outline level'),
        ('UEPOS',6,'Higher-level item in BOM structures'),
        ('MATKL',9,'IDoc material class'),
        ('MENGE',15,'Quantity'),
        ('MENEE',3,'Unit of measure'),
        ('BMNG2',15,'Quantity in price unit of measure'),
        ('PMENE',3,'Price unit of measure'),
        ('BPUMN',6,'Denominator for conversion BPRME to BME'),
        ('BPUMZ',6,'Numerator for conversion BPRME to BME'),
        ('VPREI',15,'Price (net)'),
        ('PEINH',9,'Price unit'),
        ('NETWR',18,'Item value (net)'),
        ('ANETW',18,'Absolute net value of item'),
        ('SKFBP',18,'Cash-discount-eligible amount'),
        ('CURCY',3,'Currency'),
        ('PREIS',18,'Gross price'),
        ('ACTION',3,'Action code affecting the item'),
        ('KZABS',1,'Flag for order acknowledgment requirement'),
        ('UEBTO',4,'Over-delivery tolerance'),
        ('UEBTK',1,'Unlimited over-delivery'),
        ('LBNUM',3,'Description of service area'),
        ('AUSGB',4,'Output of service area'),
        ('FRPOS',6,'Lower limit'),
        ('TOPOS',6,'Upper limit'),
        ('KTXT1',40,'Short text'),
        ('KTXT2',40,'Short text'),
        ('PERNR',8,'Personnel number'),
        ('LGART',4,'Wage type'),
        ('STELL',8,'Position'),
        ('ZWERT',18,'Total value sum segment'),
    ],
    'E2EDCA': [
        ('PARVW',3,'Partner role (e.g. sold-to party, ship-to party, ...)'),
        ('PARTN',17,'Partner number'),
        ('LIFNR',17,'Vendor number at customer'),
        ('NAME1',35,'Name 1'),
        ('NAME2',35,'Name 2'),
        ('NAME3',35,'Name 3'),
        ('NAME4',35,'Name 4'),
        ('STRAS',35,'Street and house number 1'),
        ('STRS2',35,'Street and house number 2'),
        ('PFACH',35,'PO box'),
        ('ORT01',35,'City'),
        ('COUNC',9,'County code'),
        ('PSTLZ',9,'Postal code'),
        ('PSTL2',9,'PO box postal code'),
        ('LAND1',3,'Country key'),
        ('ABLAD',35,'Unloading point'),
        ('PERNR',30,'Personnel number of contact person'),
        ('PARNR',30,'Contact person number (not personnel number)'),
        ('TELF1',25,'Telephone number 1 of contact person'),
        ('TELF2',25,'Telephone number 2 of contact person'),
        ('TELBX',25,'Telebox number'),
        ('TELFX',25,'Fax number'),
        ('TELTX',25,'Teletex number'),
        ('TELX1',25,'Telex number'),
        ('SPRAS',1,'Language key'),
        ('ANRED',15,'Title/salutation'),
        ('ORT02',35,'District'),
        ('HAUSN',6,'House number'),
        ('STOCK',6,'Floor'),
        ('REGIO',3,'Region'),
        ('PARGE',1,'Gender of partner'),
        ('ISOAL',2,'ISO code of country/region'),
        ('ISONU',2,'ISO code of country/region'),
        ('FCODE',20,'French company code'),
        ('IHREZ',30,'Your reference (from partner)'),
        ('BNAME',35,'IDoc user name'),
        ('PAORG',30,'IDoc organization code'),
        ('ORGTX',35,'IDoc organization code'),
        ('PAGRU',30,'IDoc group code'),
    ],
    'E2EDCT': [
        ('TDID',4,'Text ID'),
        ('TSSPRAS',3,'Language key'),
    ],
    'E2EDK': [
        ('ACTION',3,'Action code affecting the entire EDI message'),
        ('KZABS',1,'Flag for order acknowledgment requirement'),
        ('CURCY',3,'Currency'),
        ('HWAER',3,'EDI local currency'),
        ('WKURS',12,'Exchange rate'),
        ('ZTERM',17,'Payment terms key'),
        ('KUNDEUINR',20,'VAT registration number'),
        ('EIGENUINR',20,'VAT registration number'),
        ('BSART',4,'Document type'),
        ('BELNR',35,'IDoc document number'),
        ('NTGEW',18,'Net weight'),
        ('BRGEW',18,'Net weight'),
        ('GEWEI',3,'Weight unit'),
        ('FKART_RL',4,'Invoice list type'),
        ('ABLAD',25,'Unloading point'),
        ('BSTZD',4,'Customer purchase order number supplement'),
        ('VSART',2,'Shipping condition'),
        ('VSART_BEZ',20,'Shipping type description'),
        ('RECIPNT_NO',10,'Recipient number (for control via ALE model)'),
        ('KZAZU',1,'Order combination indicator'),
        ('AUTLF',1,'Complete delivery per order required?'),
        ('AUGRU',3,'Order reason (reason for business transaction)'),
        ('AUGRU_BEZ',40,'Description'),
        ('ABRVW',3,'Usage indicator'),
        ('ABRVW_BEZ',20,'Description'),
        ('FKTYP',1,'Billing type'),
        ('LIFSK',2,'Delivery block (document header)'),
        ('LIFSK_BEZ',20,'Description'),
        ('EMPST',25,'Receiving point'),
        ('ABTNR',4,'Department number'),
        ('DELCO',3,'Agreed delivery time'),
        ('WKURS_M',12,'Quantity-quoted rate in an IDoc segment'),
        ('LANDTX',3,'Tax departure country/region'),
        ('STCEG_L',3,'Country/region of VAT registration number'),
    ],
    'E2EDKA': [
        ('PARVW',3,'Partner role (e.g. sold-to party, ship-to party, ...)'),
        ('PARTN',17,'Partner number'),
        ('LIFNR',17,'Vendor number at customer'),
        ('NAME1',35,'Name 1'),
        ('NAME2',35,'Name 2'),
        ('NAME3',35,'Name 3'),
        ('NAME4',35,'Name 4'),
        ('STRAS',35,'Street and house number 1'),
        ('STRS2',35,'Street and house number 2'),
        ('PFACH',35,'PO box'),
        ('ORT01',35,'City'),
        ('COUNC',9,'County code'),
        ('PSTLZ',9,'Postal code'),
        ('PSTL2',9,'Postal code of PO box'),
        ('LAND1',3,'Country key'),
        ('ABLAD',35,'Unloading point'),
        ('PERNR',30,'Personnel number of contact person'),
        ('PARNR',30,'Contact person number (not personnel number)'),
        ('TELF1',25,'Phone number 1 of contact person'),
        ('TELF2',25,'Phone number 2 of contact person'),
        ('TELBX',25,'Telebox number'),
        ('TELFX',25,'Fax number'),
        ('TELTX',25,'Teletex number'),
        ('TELX1',25,'Telex number'),
        ('SPRAS',1,'Language key'),
        ('ANRED',15,'Salutation'),
        ('ORT02',35,'District'),
        ('HAUSN',6,'House number'),
        ('STOCK',6,'Floor'),
        ('REGIO',3,'Region'),
        ('PARGE',1,'Gender of partner'),
        ('ISOAL',2,'ISO code of country/region'),
        ('ISONU',2,'ISO code of country/region'),
        ('FCODE',20,'Company code France'),
        ('IHREZ',30,'Your reference (from partner)'),
        ('BNAME',35,'IDoc user name'),
        ('PAORG',30,'IDoc organization code'),
        ('ORGTX',35,'IDoc organization code'),
        ('PAGRU',30,'IDoc group code'),
        ('KNREF',30,'Customer-specific designation of business partner (plant, st'),
        ('ILNNR',70,'Character field of length 70'),
        ('PFORT',35,'City of PO box'),
        ('SPRAS_ISO',2,'2-character SAP language code'),
        ('TITLE',15,'Salutation'),
    ],
    'E2EDKT': [
        ('TDID',4,'Text ID'),
        ('TSSPRAS',3,'Language key'),
        ('TSSPRAS_ISO',2,'2-character SAP language code'),
        ('TDOBJECT',10,'Texts: application object'),
        ('TDOBNAME',70,'Name'),
    ],
    'E2EDL': [
        ('EXIDV',20,'External handling unit identification'),
        ('TARAG',17,'Tare weight of handling unit'),
        ('GWEIT',3,'Tare weight unit'),
        ('BRGEW',17,'Total weight of handling unit'),
        ('NTGEW',17,'Load weight of handling unit'),
        ('MAGEW',17,'Permissible load weight of handling unit'),
        ('GWEIM',3,'Weight unit'),
        ('BTVOL',17,'Total volume of handling unit'),
        ('NTVOL',17,'Load volume of handling unit'),
        ('MAVOL',17,'Permissible load volume of handling unit'),
        ('VOLEM',3,'Volume unit'),
        ('TAVOL',17,'Tare volume of handling unit'),
        ('VOLET',3,'Tare volume unit'),
        ('VEGR2',5,'Handling unit group 2 (freely definable)'),
        ('VEGR1',5,'Handling unit group 1 (freely definable)'),
        ('VEGR3',5,'Handling unit group 3 (freely definable)'),
        ('VHILM',18,'Packaging materials'),
        ('VEGR4',5,'Handling unit group 4 (freely definable)'),
        ('LAENG',15,'Length'),
        ('VEGR5',5,'Handling unit group 5 (freely definable)'),
        ('BREIT',15,'Width'),
        ('HOEHE',15,'Height'),
        ('MEABM',3,'Unit for length/width/height'),
        ('INHALT',40,'Description of handling unit contents'),
        ('VHART',4,'Packaging material type'),
        ('MAGRV',4,'Material group for packaging'),
        ('LADLG',8,'Loading length in loading length unit'),
        ('LADEH',3,'Loading length unit'),
        ('FARZT',4,'Travel time'),
        ('FAREH',3,'Travel time unit'),
        ('ENTFE',8,'Distance traveled'),
        ('EHENT',3,'Distance unit'),
        ('VELTP',1,'Packaging material type'),
        ('EXIDV2',20,'2nd external identification of handling units'),
        ('LANDT',3,'Country/region of means of transport'),
        ('LANDF',3,'Driver nationality'),
        ('NAMEF',35,'Driver name'),
        ('NAMBE',35,'Co-driver name'),
        ('VHILM_KU',22,'Customer material'),
        ('VEBEZ',40,'Packaging material description'),
        ('SMGKN',1,'SMG identifier for goods tag'),
        ('KDMAT35',35,'Partner packaging material (customer / vendor)'),
        ('SORTL',10,'Sort field'),
        ('ERNAM',12,'Name of the clerk who added the object'),
        ('GEWFX',1,'Fixed weights and volumes'),
        ('ERLKZ',1,'Status (currently without functionality)'),
        ('EXIDA',1,'Type of external handling unit identification'),
        ('MOVE_STATUS',4,'Status of handling unit'),
        ('PACKVORSCHR',22,'Text string 22 characters'),
        ('PACKVORSCHR_ST',1,'Single-character indicator'),
        ('LABELTYP',1,'Indicator: do not print external shipping label'),
        ('ZUL_AUFL',17,'Field of length 17'),
        ('VHILM_EXTERNAL',40,'Shipping material'),
        ('VHILM_VERSION',10,'Version number for field VHILM'),
        ('VHILM_GUID',32,'External GUID for field VHILM'),
        ('KDMAT35_EXTERNAL',40,'Long material number (future development) for field KDMAT'),
        ('KDMAT35_VERSION',10,'Version number (future development) for field KDMAT35'),
        ('KDMAT35_GUID',32,'External GUID (future development) for field KDMAT35'),
        ('VHILM_KU_EXTERNAL',40,'Long material number (future development) for field VHILM'),
        ('VHILM_KU_VERSION',10,'Version number (future development) for field VHILM_KU'),
        ('VHILM_KU_GUID',32,'External GUID (future development) for field VHILM_KU'),
        ('VHILM_LONG',40,'Packaging materials'),
    ],
    'E2EDP': [
        ('POSEX',6,'Item number'),
        ('ACTION',3,'Action code affecting the item'),
        ('PSTYP',1,'Item type'),
        ('KZABS',1,'Flag for order acknowledgment requirement'),
        ('MENGE',15,'Quantity'),
        ('MENEE',3,'Unit of measure'),
        ('BMNG2',15,'Quantity in price unit of measure'),
        ('PMENE',3,'Price unit of measure'),
        ('ABFTZ',7,'Cumulative quantity reconciliation number'),
        ('VPREI',15,'Price (net)'),
        ('PEINH',9,'Price unit'),
        ('NETWR',18,'Item value (net)'),
        ('ANETW',18,'Absolute net value of item'),
        ('SKFBP',18,'Cash-discount-eligible amount'),
        ('NTGEW',18,'Net weight'),
        ('GEWEI',3,'Weight unit'),
        ('EINKZ',1,'Flag: more than one schedule line per item'),
        ('CURCY',3,'Currency'),
        ('PREIS',18,'Gross price'),
        ('MATKL',9,'IDoc material class'),
        ('UEPOS',6,'Higher-level item in BOM structures'),
        ('GRKOR',3,'Delivery group (items delivered together)'),
        ('EVERS',7,'Shipping instructions'),
        ('BPUMN',6,'Denominator for conversion BPRME to BME'),
        ('BPUMZ',6,'Numerator for conversion BPRME to BME'),
        ('ABGRU',2,'Rejection reason for sales documents'),
        ('ABGRT',40,'Description'),
        ('ANTLF',1,'Maximum number of partial deliveries allowed per item'),
        ('FIXMG',1,'Delivery date and quantity fixed'),
        ('KZAZU',1,'Order combination indicator'),
        ('BRGEW',18,'Total weight'),
        ('PSTYV',4,'Item type in SD document'),
        ('EMPST',25,'Receiving point'),
        ('ABTNR',4,'Department number'),
        ('ABRVW',3,'Usage indicator'),
        ('WERKS',4,'Plant'),
        ('LPRIO',2,'Delivery priority'),
        ('LPRIO_BEZ',20,'Description'),
        ('ROUTE',6,'Route'),
        ('ROUTE_BEZ',40,'Description'),
        ('LGORT',4,'Storage location'),
        ('VSTEL',4,'Shipping point / receiving point'),
        ('DELCO',3,'Agreed delivery time'),
        ('MATNR',35,'IDoc material identification'),
        ('VALTG',2,'Additional value days'),
        ('HIPOS',6,'Higher-level item in an item hierarchy'),
        ('HIEVW',1,'Usage of hierarchy item'),
        ('POSGUID',22,'ATP: encoding of DELNR and DELPS'),
        ('MATNR_EXTERNAL',40,'Material number'),
        ('MATNR_VERSION',10,'Version number for field MATNR'),
        ('MATNR_GUID',32,'External GUID for field MATNR'),
        ('IUID_RELEVANT',1,'IUID-relevant'),
        ('SGT_RCAT',16,'Requirements segment'),
        ('SGT_SCAT',16,'Stock segment'),
        ('HANDOVERLOC',10,'Location for physical handover of goods'),
        ('MATNR_LONG',40,'Material number'),
        ('REQ_SEG_LONG',40,'Requirements segment'),
        ('STK_SEG_LONG',40,'Stock segment'),
        ('EXPECTED_VALUE',31,'Currency amount for BAPIs (with 9 decimal places)'),
        ('LIMIT_AMOUNT',31,'Currency amount for BAPIs (with 9 decimal places)'),
    ],
    'E2EDPA': [
        ('PARVW',3,'Partner role (e.g. sold-to party, ship-to party, ...)'),
        ('PARTN',17,'Partner number'),
        ('LIFNR',17,'Vendor number at customer'),
        ('NAME1',35,'Name 1'),
        ('NAME2',35,'Name 2'),
        ('NAME3',35,'Name 3'),
        ('NAME4',35,'Name 4'),
        ('STRAS',35,'Street and house number 1'),
        ('STRS2',35,'Street and house number 2'),
        ('PFACH',35,'PO box'),
        ('ORT01',35,'City'),
        ('COUNC',9,'County code'),
        ('PSTLZ',9,'Postal code'),
        ('PSTL2',9,'PO box postal code'),
        ('LAND1',3,'Country key'),
        ('ABLAD',35,'Unloading point'),
        ('PERNR',30,'Personnel number of contact person'),
        ('PARNR',30,'Contact person number (not personnel number)'),
        ('TELF1',25,'Telephone number 1 of contact person'),
        ('TELF2',25,'Telephone number 2 of contact person'),
        ('TELBX',25,'Telebox number'),
        ('TELFX',25,'Fax number'),
        ('TELTX',25,'Teletex number'),
        ('TELX1',25,'Telex number'),
        ('SPRAS',1,'Language key'),
        ('ANRED',15,'Title/salutation'),
        ('ORT02',35,'District'),
        ('HAUSN',6,'House number'),
        ('STOCK',6,'Floor'),
        ('REGIO',3,'Region'),
        ('PARGE',1,'Gender of partner'),
        ('ISOAL',2,'ISO code of country/region'),
        ('ISONU',2,'ISO code of country/region'),
        ('FCODE',20,'French company code'),
        ('IHREZ',30,'Your reference (from partner)'),
        ('BNAME',35,'IDoc user name'),
        ('PAORG',30,'IDoc organization code'),
        ('ORGTX',35,'IDoc organization code'),
        ('PAGRU',30,'IDoc group code'),
        ('KNREF',30,'Customer-specific description of business partner (plant, wa'),
        ('ILNNR',70,'Character field of length 70'),
        ('PFORT',35,'City of PO box'),
        ('SPRAS_ISO',2,'2-character SAP language code'),
        ('TITLE',15,'Title/salutation'),
    ],
    'E2EDPAD': [
        ('QUALF',3,'IDoc object identification (A&D)'),
        ('ICC',2,'Interchangeability indicator'),
        ('MOI',4,'Type indicator'),
        ('PRI',3,'Order priority'),
        ('ACN',5,'Aircraft registration number'),
        ('ACN10',10,'Aircraft registration number'),
        ('DNS',1,'Do not replace'),
    ],
    'E2EDPS': [
        ('KSTBM',17,'Condition scale quantity'),
        ('KBETR',13,'Condition amount or percentage'),
    ],
    'E2EDPT': [
        ('TDID',4,'Text ID'),
        ('TSSPRAS',3,'Language key'),
        ('TSSPRAS_ISO',2,'2-character SAP language code'),
    ],
    'E2EDS': [
        ('SUMID',3,'Sum segment qualifier for delivery note'),
        ('SUMME',18,'Total value sum segment'),
        ('SUNIT',3,'Unit of total value sum segment delivery note'),
        ('WAERQ',3,'Currency'),
    ],
    'E2IDOCENHANCEMENT': [
        ('IDENTIFIER',30,'Data container for IDoc type extensions'),
        ('DATA',970,'Data container for IDoc extension'),
    ],
    'E2TXTH': [
        ('FUNCTION',3,'Function (for transferred text)'),
        ('TDOBJECT',10,'Texts: application object'),
        ('TDOBNAME',70,'Name'),
        ('TDID',4,'Text ID'),
        ('TDSPRAS',1,'Language'),
        ('TDTEXTTYPE',6,'SAPscript: format of a text'),
        ('LANGUA_ISO',2,'Language indicator'),
    ],
    'E2TXTP': [
        ('TDFORMAT',2,'Format column'),
        ('TDLINE',132,'Text line'),
    ],
    'ZE1EDKA': [
        ('STR_SUPPL2',40,'Street 3'),
        ('STR_SUPPL3',40,'Street 4'),
        ('LOCATION',40,'Street 5'),
    ],
    'ZE1EDKEMAIL': [
        ('QUALF',3,'Three-character field for IDocs'),
        ('EMAIL',241,'E-mail address'),
    ],
    'ZE1EDP': [
        ('BEDNR',10,'Requirements number'),
    ],
    'ZITEM_EXTN': [
        ('BACK_ORDER_QTY',15,'BACK_ORDER_QTY'),
        ('BACK_ORDER_UOM',3,'Unit of measure'),
        ('NEW_MATNR',18,'Material number'),
        ('PRO_NO',132,'Text line'),
        ('LINE_STATUS',2,'LINE_STATUS'),
        ('QTY_CHANGED',15,'QTY_CHANGED'),
        ('UOM',3,'UOM'),
        ('STATUS_DATE',8,'STATUS_DATE'),
        ('STATUS_REFERENCE',45,'STATUS_REFERENCE'),
    ],
    'ZITMACK': [
        ('STATUS',2,'STATUS'),
        ('QTY',15,'QTY'),
        ('UOM',3,'UOM'),
        ('DATE',8,'DATE'),
        ('REF',45,'REF'),
    ],

    # ── DELFOR02-specific segments ─────────────────────────────────────────
    'E2EDK09': [
        ('VTRNR',35,'Contract number'),
        ('BSTDK',8,'Customer reference date'),
        ('LABNK',17,'Current release key of the customer'),
        ('ZEICH',70,"Customer's reference mark"),
        ('BSTZD',70,'Account assignment'),
        ('ABRVW',3,'Usage indicator'),
        ('KSTAT',2,'Converter status'),
        ('KTEXT',70,'Error text from converter'),
        ('ABNRA',17,'Current release key of the customer'),
        ('ABNRD',8,'Date of the release'),
        ('KTEXT_V',40,'Search term assortment'),
        ('USR01',35,'Customer reserve: additional data field 1'),
        ('USR02',35,'Customer reserve: additional data field 2'),
        ('USR03',35,'Customer reserve: additional data field 3'),
        ('USR04',10,'Customer reserve: additional data field 4'),
        ('USR05',10,'Customer reserve: additional data field 5'),
        ('CYEFZ',15,'Cumulative quantity reached at the date of reset'),
        ('CYDAT',8,'Date of reset of the inbound cumulative quantity'),
        ('MFLAUF',3,'Material release lead time'),
        ('MFEIN',1,'Material release - lead time unit'),
        ('FFLAUF',3,'Production release lead time'),
        ('FFEIN',1,'Production release - lead time unit'),
    ],
    'E2EDK10': [('DUMMY',1,'Dummy function of length 1')],
    'E2EDK11': [
        ('TDNAME',4,'Text role EDI'),('SPRAS',3,'Language key'),
        ('TXTLF',2,'Continuation text (sequential number)'),
        ('TXT01',70,'Sequential text line 1'),('TXT02',70,'Sequential text line 2'),
        ('TXT03',70,'Sequential text line 3'),('TXT04',70,'Sequential text line 4'),
        ('TXT05',70,'Sequential text line 5'),('TXT06',70,'Sequential text line 6'),
        ('TXT07',70,'Sequential text line 7'),('TXT08',70,'Sequential text line 8'),
        ('TXT09',70,'Sequential text line 9'),('TXT10',70,'Sequential text line 10'),
        ('TXT11',70,'Sequential text line 11'),('TXT12',70,'Sequential text line 12'),
        ('TXT13',70,'Sequential text line 13'),('TXT14',70,'Sequential text line 14'),
        ('TSSPRAS_ISO',2,'2-character SAP language code'),
    ],
    'E2EDP10': [
        ('IDNKD',35,'Customer material number'),
        ('ARKTX',35,'Item short text EDI'),
        ('VRKME',3,'Sales unit of measure'),
        ('KWERK',20,"Customer's plant"),
        ('KLGOR',7,'EDI storage location of customer'),
        ('DFABL',17,"Customer's unloading point"),
        ('VBRST',14,"Customer's consumption point"),
        ('BELNR',35,'Delivery note number of last EDI receipt'),
        ('LFIMG',15,'Delivery note quantity of last EDI schedule line'),
        ('VEMNG',15,'Received quantity of last schedule line'),
        ('LIDTL',8,'Date of last delivery posted by customer'),
        ('LIFST',1,'Delivery status key'),
        ('KRITB',1,'Critical stock key'),
        ('ABHOR',8,'Fine release horizon'),
        ('ABDAT',8,'Reconciliation date for reconciliation cumulative quantity'),
        ('FZDIF',15,'Cumulative quantity difference EDI'),
        ('AKUBM',15,'Current cumulative order quantity'),
        ('AKUEM',15,'Current cumulative received quantity'),
        ('DESRE',1,'Design revision number'),
        ('ECHNO',1,'Engineering Change Number'),
        ('ECHDT',8,'Engineering Change Date'),
        ('ABFDE',8,'Production release end'),
        ('MFADT',15,'Production release cumulative requirement'),
        ('ABMDE',8,'Material release end'),
        ('FPSDN',22,'FPSD number'),
        ('LABKY',1,'Delivery release key'),
        ('TSTKY',2,'Part status key'),
        ('DOKPF',1,'Documentation requirement for inspection results'),
        ('KZAUS',1,'Phase-out indicator'),
        ('ABRAB',8,'Scheduling agreement release valid from'),
        ('ABRBI',8,'Scheduling agreement release valid to'),
        ('SCREL',2,'Schedule Release'),
        ('NULDT',8,'Reset date'),
        ('TEART',1,'Dispatch / arrival date indicator'),
        ('SOLLFZ',15,'Target cumulative quantity'),
        ('SOLLDT',8,'Target date for target cumulative quantity'),
        ('IDNLF',35,'Vendor material number'),
        ('BSTDK',8,'Customer reference date'),
        ('LABNK',17,'Current release key of the customer'),
        ('ABNRA',17,'Current release key of the customer'),
        ('ABNRD',8,'Date of the release'),
        ('VTRNR',35,'Contract number'),
        ('MFRFZ',15,'Material release cumulative requirement'),
        ('POSEX',6,'Item number of the underlying purchase order'),
        ('IDNKD_EXTERNAL',40,'Long material number (future development) for field IDNKD'),
        ('IDNKD_GUID',32,'External GUID (future development) for field IDNKD'),
        ('IDNKD_VERSION',10,'Version number (future development) for field IDNKD'),
    ],
    'E2EDP11': [('DUMMY',1,'Dummy function of length 1')],
    'E2EDP14': [
        ('PCKAR',35,'Packaging type'),
        ('PCKNR',35,"Customer's package number"),
        ('ANZPK',15,'Number of packages'),
        ('PCKSL',2,'Packaging type key'),
        ('ANZAR',15,'Number of items per package'),
        ('MAZAR',3,'Unit of measure for items/package'),
        ('MMPCK',3,'Unit of measure for L/W/H of package'),
        ('LPACK',5,'Length of package'),
        ('BPACK',5,'Width of package'),
        ('HPACK',5,'Height of package'),
        ('MGPCK',3,'Unit of measure for weight/package'),
        ('GPACK',11,'Gross weight of package'),
        ('NPACK',11,'Net weight of package'),
        ('VPACK',11,'Volume of package'),
        ('GDATV',8,'Packaging validity date'),
    ],
    'E2EDP15': [
        ('TDNAME',4,'Text role EDI'),('SPRAS',3,'Language key'),
        ('TXTLF',2,'Continuation text (sequential number)'),
        ('TXT01',70,'Sequential text line 1'),('TXT02',70,'Sequential text line 2'),
        ('TXT03',70,'Sequential text line 3'),('TXT04',70,'Sequential text line 4'),
        ('TXT05',70,'Sequential text line 5'),('TXT06',70,'Sequential text line 6'),
        ('TXT07',70,'Sequential text line 7'),('TXT08',70,'Sequential text line 8'),
        ('TXT09',70,'Sequential text line 9'),('TXT10',70,'Sequential text line 10'),
        ('TXT11',70,'Sequential text line 11'),('TXT12',70,'Sequential text line 12'),
        ('TXT13',70,'Sequential text line 13'),('TXT14',70,'Sequential text line 14'),
        ('TSSPRAS_ISO',2,'2-character SAP language code'),
    ],
    'E2EDP16': [
        ('ETTYP',1,'Schedule line type EDI'),
        ('PRGRS',1,'Date type of the date'),
        ('EDATUV',8,'Schedule line date from'),
        ('EZEIT',4,'Release time'),
        ('EDATUB',8,'Schedule line date to'),
        ('ETVTF',2,'Schedule line distribution function'),
        ('WMENG',15,'Release quantity'),
        ('FZABR',15,'Inbound cumulative quantity of schedule line'),
        ('BSTAS',1,'Requirement status key'),
        ('WDATUV',8,'Planned goods issue date from'),
        ('WZEIT',4,'Planned goods issue time'),
        ('WDATUB',8,'Planned goods issue date to'),
        ('LDATUV',8,'Planned goods receipt date from'),
        ('LZEIT',4,'Planned goods issue time'),
        ('LDATUB',8,'Planned goods receipt date to'),
        ('BSTGRU',3,'Order reason'),
        ('TRANSPTP',10,'Means of transport'),
        ('TRANSPDESC',40,'Description of means of transport'),
    ],
    'E2EDP36': [
        ('LIFST',1,'Delivery status key'),
        ('BELNR',35,'Delivery note number of last EDI receipt'),
        ('BELDAT',8,'Date of delivery note'),
        ('BELTIM',6,'Time of delivery note'),
        ('LFIMG',15,'Delivery note quantity of last EDI schedule line'),
        ('VRKME',3,'Sales unit of measure'),
        ('VEMNG',15,'Received quantity of last schedule line'),
        ('LIKME',3,'Order unit of measure'),
        ('LIDTL',8,'Date of goods receipt'),
        ('LIDTIM',6,'Time of goods receipt'),
    ],

}

SEG_DESC = {
    'EDI_DC40':    'Control Record – nagłówek IDoc',
    'E2EDK01':     'Nagłówek zamówienia (waluta, warunki płatności, typ)',
    'ZE1EDKEMAIL': 'Adres e-mail (rozszerzenie Z)',
    'E2EDK14':     'Jednostki organizacyjne (zakład, org. zakupów)',
    'E2EDK03':     'Daty dokumentu (zamówienia, dostawy)',
    'E2EDKA1':     'Dane adresowe partnera (AG/LF/WE)',
    'E2EDK02':     'Numer referencyjny dokumentu',
    'E2EDK17':     'Warunki dostawy / Incoterms',
    'E2EDK18':     'Warunki płatności',
    'E2EDK35':     'Opakowania / dodatkowe dane',
    'E2EDKT1':     'Nagłówek tekstu (nagłówkowy)',
    'E2EDKT2':     'Linia tekstu (nagłówkowy)',
    'E2EDP01':     'Pozycja zamówienia (materiał, ilość, cena)',
    'E2EDP20':     'Harmonogram dostaw pozycji',
    'E2EDP19':     'Identyfikacja materiału (nr SAP, nr dostawcy)',
    'E2EDP17':     'Warunki dostawy pozycji',
    'E2EDPT1':     'Nagłówek tekstu (pozycja)',
    'E2EDPT2':     'Linia tekstu (pozycja)',
    'E2EDS01':     'Suma kontrolna dokumentu',
}

# ── Helpers ────────────────────────────────────────────────────────────────
def get_key(seg_name):
    # 1. Exact full name
    if seg_name in SEGMENT_FIELDS:
        return seg_name
    # 2. Strip 3-digit version suffix (E2EDP16002 -> E2EDP16)
    m = re.match(r'^(.*?)(\d{3})$', seg_name)
    if m:
        base3 = m.group(1)
        if base3 in SEGMENT_FIELDS:
            return base3
    # 3. Strip all trailing digits (E2EDKA1003 -> E2EDKA)
    base_all = re.sub(r'\d+$', '', seg_name)
    if base_all in SEGMENT_FIELDS:
        return base_all
    # 4. Longest-key-first prefix match
    for k in sorted(SEGMENT_FIELDS.keys(), key=len, reverse=True):
        if seg_name.startswith(k):
            return k
    return None

def parse_flat(filepath):
    rows = []
    with open(filepath, 'r', encoding='latin-1') as f:
        for line in f:
            line = line.rstrip('\n')
            if not line.strip():
                continue
            seg_name = line[:30].strip()
            try:
                hlevel = int(line[61:63].strip())
            except (ValueError, IndexError):
                hlevel = 0
            rows.append((seg_name, hlevel, line[63:]))
    return rows

def extract_fields(seg_name, data):
    key = get_key(seg_name)
    if not key:
        return [('(raw)', len(data), '(unknown segment)', data.strip())]
    result, pos = [], 0
    for fname, flen, fdesc in SEGMENT_FIELDS[key]:
        val = data[pos:pos+flen] if pos < len(data) else ''
        result.append((fname, flen, fdesc, val.strip()))
        pos += flen
    return result

# ── Field value lookups ────────────────────────────────────────────────────
PARVW_DESC = {
    'AG': 'Zamawiający (sold-to party)',
    'LF': 'Dostawca (vendor)',
    'WE': 'Odbiorca dostawy (ship-to party)',
    'RE': 'Płatnik (bill-to party)',
    'RG': 'Płatnik faktury (payer)',
    'EK': 'Kupiec (buyer)',
    'AP': 'Osoba kontaktowa (contact person)',
    'VN': 'Dostawca alternatywny (alt. vendor)',
    'LS': 'Nadawca logiczny (logical system)',
    'PE': 'Osoba odpowiedzialna (person responsible)',
    'AB': 'Adres zamówienia (ordering address)',
    'SP': 'Spedytor (freight forwarder)',
    'SH': 'Nadawca (shipper)',
    'CN': 'Odbiorca (consignee)',
    'GS': 'Nadawca towarów (goods supplier)',
    'Z1': 'Partner Z1 (custom)',
    'Z2': 'Partner Z2 (custom)',
}

QUALF_EDK14 = {
    '001': 'Jednostka gospodarcza (company code)',
    '006': 'Oddział (business area)',
    '007': 'Centrum zysku (profit center)',
    '008': 'Controlling area',
    '009': 'Obszar sprzedaży (sales area)',
    '010': 'Kanał dystrybucji (distribution channel)',
    '011': 'Zakład (plant)',
    '012': 'Miejsce składowania (storage location)',
    '013': 'Typ zamówienia zakupu (purchase order type)',
    '014': 'Organizacja zakupów (purchasing org.)',
    '015': 'Grupa zakupów (purchasing group)',
    '016': 'Obszar sprzedaży SD',
    '017': 'Dywizja (division)',
    '018': 'Organizacja sprzedaży (sales org.)',
    '019': 'Biuro sprzedaży (sales office)',
    '020': 'Grupa sprzedaży (sales group)',
}

QUALF_EDK03 = {
    '001': 'Data żądanej dostawy (requested delivery)',
    '002': 'Data zamówienia (purchase order date)',
    '004': 'Data dostawy (delivery date)',
    '007': 'Data oferty (quotation date)',
    '008': 'Data ważności oferty',
    '009': 'Data potwierdzenia zamówienia',
    '010': 'Data wystawienia faktury',
    '011': 'Data dostawy od',
    '012': 'Data dostawy do',
    '013': 'Data płatności',
    '014': 'Data kontraktu',
    '015': 'Data harmonogramu',
    '017': 'Data wysyłki',
    '018': 'Data przyjęcia towaru (GR date)',
    '019': 'Data dokumentu',
    '020': 'Data wejścia w życie',
}

QUALF_EDK02 = {
    '001': 'Numer zamówienia klienta (customer PO)',
    '002': 'Numer potwierdzenia zamówienia',
    '003': 'Numer kontraktu',
    '004': 'Numer zapytania ofertowego (RFQ)',
    '005': 'Numer oferty (quotation)',
    '006': 'Numer dostawy',
    '007': 'Numer faktury',
    '008': 'Numer listu przewozowego (bill of lading)',
    '009': 'Numer AWB / dokumentu spedycyjnego',
    '010': 'Numer planu dostaw (scheduling agreement)',
    '011': 'Numer zamówienia zakupu (PO number)',
    '012': 'Numer partii (batch)',
    '013': 'Numer projektu',
    '015': 'Numer referencyjny EDI',
}

QUALF_EDK17 = {
    '001': 'Kod Incoterms (np. FCA, EXW, DAP)',
    '002': 'Miejsce dostawy Incoterms',
    'VER': 'Uwaga / warunek dodatkowy (remark)',
}

QUALF_EDK18 = {
    '001': 'Termin płatności netto',
    '002': 'Skonto 1 (cash discount 1)',
    '003': 'Skonto 2 (cash discount 2)',
    '004': 'Termin płatności 4',
}

QUALF_EDK35 = {
    'ZZE': 'Opakowanie zwrotne (returnable)',
    'ZZN': 'Opakowanie jednorazowe (non-returnable)',
    'AAA': 'Standardowe opakowanie',
    'LPP': 'Opakowanie paletowe',
}

QUALF_EDP19 = {
    '001': 'Numer materiału SAP (internal)',
    '002': 'Numer EAN / GTIN (barcode)',
    '003': 'Numer materiału klienta',
    '004': 'Numer materiału dostawcy',
    '005': 'Numer rysunku (drawing)',
    '006': 'Numer rewizji (revision)',
    '007': 'Numer seryjny',
    '008': 'Numer partii (batch)',
    '009': 'Numer katalogowy',
    '010': 'Numer zamówienia zakupu',
    '011': 'Numer umowy ramowej',
    'HIS': 'Numer historyczny',
    'MAN': 'Numer producenta',
}

QUALF_ZEMAIL = {
    'Z01': 'E-mail główny (primary)',
    'Z02': 'E-mail osoby kontaktowej (contact person)',
    'Z03': 'E-mail do faktur (invoices)',
    'Z04': 'E-mail do zamówień (orders)',
    'Z05': 'E-mail dodatkowy (additional)',
}

FIELD_LOOKUPS = {
    ('E2EDKA',  'PARVW'): PARVW_DESC,
    ('E2EDKA1', 'PARVW'): PARVW_DESC,
    ('E2EDPA',  'PARVW'): PARVW_DESC,
    ('E2EDK09', 'PARVW'): PARVW_DESC,
    ('E2EDK14', 'QUALF'): QUALF_EDK14,
    ('E2EDK03', 'QUALF'): QUALF_EDK03,
    ('E2EDK02', 'QUALF'): QUALF_EDK02,
    ('E2EDK17', 'QUALF'): QUALF_EDK17,
    ('E2EDP17', 'QUALF'): QUALF_EDK17,
    ('E2EDK18', 'QUALF'): QUALF_EDK18,
    ('E2EDK35', 'QUALF'): QUALF_EDK35,
    ('E2EDP19', 'QUALF'): QUALF_EDP19,
    ('ZE1EDKEMAIL', 'QUALF'): QUALF_ZEMAIL,
}


def field_meaning(seg_name, fname, val):
    """Return human-readable meaning for coded field values, or empty string."""
    key = get_key(seg_name)
    lookup = FIELD_LOOKUPS.get((key, fname))
    if lookup and val:
        return lookup.get(val.strip(), '')
    return ''

# ── Styles ─────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill('solid', start_color='1F4E79')
SEG_FILL  = PatternFill('solid', start_color='2E75B6')
ALT_FILL  = PatternFill('solid', start_color='EBF3FB')
WHT_FILL  = PatternFill('solid', start_color='FFFFFF')
EMPTY_FILL= PatternFill('solid', start_color='FFF2CC')
HDR_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
SEG_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
NORM_FONT = Font(name='Arial', size=10)
thin      = Side(style='thin', color='BFBFBF')
BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, cols_vals):
    for col, val in enumerate(cols_vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 28

def cell(ws, r, col, val, fill=None, font=None, align=None):
    c = ws.cell(row=r, column=col, value=val)
    c.font = font or NORM_FONT
    if fill: c.fill = fill
    c.alignment = align or Alignment(vertical='center')
    c.border = BORDER
    return c

# ── Main builder ───────────────────────────────────────────────────────────
def build_excel(idoc_rows, out_path):
    wb = openpyxl.Workbook()

    # Sheet 1 – full breakdown
    ws = wb.active
    ws.title = 'IDoc - Pola i wartości'
    hdr(ws, 1, ['Segment','Wyst.','Nazwa pola (IDoc)','Dług.','Opis pola','Wartość','Znaczenie'])

    seg_counter, seg_hlevel, seg_order, r = {}, {}, {}, 2
    order_idx = 0
    for seg_name, hlevel, data in idoc_rows:
        seg_counter[seg_name] = seg_counter.get(seg_name, 0) + 1
        if seg_name not in seg_hlevel:
            seg_hlevel[seg_name] = hlevel
            seg_order[seg_name] = order_idx
            order_idx += 1
        occ = seg_counter[seg_name]
        indent = '  ' * hlevel
        fields = extract_fields(seg_name, data)
        first = True
        for fname, flen, fdesc, val in fields:
            fill = ALT_FILL if r % 2 == 0 else WHT_FILL
            if not val: fill = EMPTY_FILL
            seg_display = (indent + seg_name) if first else ''
            cell(ws, r, 1, seg_display, SEG_FILL if first else fill, SEG_FONT if first else NORM_FONT,
                 Alignment(horizontal='left', vertical='center') if first else Alignment(vertical='center'))
            cell(ws, r, 2, occ if first else '', SEG_FILL if first else fill,
                 SEG_FONT if first else NORM_FONT,
                 Alignment(horizontal='center', vertical='center'))
            for col, v in enumerate([fname, flen, fdesc, val], 3):
                cell(ws, r, col, v, fill, align=Alignment(vertical='center', wrap_text=(col==6)))
            # col 7: meaning (only for fields that have a lookup)
            meaning = field_meaning(seg_name, fname, val)
            meaning_fill = PatternFill('solid', start_color='E2EFDA') if meaning else fill
            cell(ws, r, 7, meaning, meaning_fill,
                 Font(name='Arial', size=10, italic=bool(meaning)),
                 Alignment(vertical='center'))
            first = False; r += 1

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 7
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 7
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 35
    ws.freeze_panes = 'A2'

    # Sheet 2 – summary (sorted by document order to reflect hierarchy)
    ws2 = wb.create_sheet('Podsumowanie')
    ws2.merge_cells('A1:D1')
    ws2['A1'] = 'IDoc – Podsumowanie segmentów'
    ws2['A1'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')
    ws2['A1'].alignment = Alignment(horizontal='center')
    hdr(ws2, 3, ['Segment (wcięcie = poziom zagnieżdżenia)','Liczba wystąpień','Opis segmentu','Poziom'])
    r2 = 4
    for seg, cnt in sorted(seg_counter.items(), key=lambda x: seg_order.get(x[0], 9999)):
        key = get_key(seg)
        desc = SEG_DESC.get(seg, SEG_DESC.get(key, ''))
        lvl = seg_hlevel.get(seg, 0)
        fill = ALT_FILL if r2 % 2 == 0 else WHT_FILL
        cell(ws2, r2, 1, '  ' * lvl + seg, fill, align=Alignment(horizontal='left', vertical='center'))
        cell(ws2, r2, 2, cnt, fill, align=Alignment(horizontal='center', vertical='center'))
        cell(ws2, r2, 3, desc, fill)
        cell(ws2, r2, 4, lvl, fill, align=Alignment(horizontal='center', vertical='center'))
        r2 += 1
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 50
    ws2.column_dimensions['D'].width = 10
    ws2.freeze_panes = 'A4'

    # Sheet 3 – positions (only if E2EDP01 present)
    has_positions = any(row[0].startswith('E2EDP01') for row in idoc_rows)
    if has_positions:
        ws3 = wb.create_sheet('Pozycje zamówienia')
        ws3.merge_cells('A1:I1')
        ws3['A1'] = 'Pozycje zamówienia – kluczowe dane'
        ws3['A1'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')
        ws3['A1'].alignment = Alignment(horizontal='center')
        hdr(ws3, 3, ['Poz.','Materiał (IDTNR)','Opis materiału','Ilość','JM','Cena jedn.','Wartość netto','Data dostawy','Zakład'])

        positions, delivery_dates, materials, cur = {}, {}, {}, None
        for seg_name, hlevel, data in idoc_rows:
            if seg_name.startswith('E2EDP01'):
                posex = data[0:6].strip(); cur = posex
                positions[posex] = {
                    'menge': data[10:25].strip(), 'menee': data[25:28].strip(),
                    'vprei': data[35:50].strip(), 'netwr': data[59:74].strip(),
                    'werks': data[193:197].strip(),
                }
            elif seg_name.startswith('E2EDP20') and cur:
                d = data[30:38].strip()
                if len(d) == 8: d = f"{d[:4]}-{d[4:6]}-{d[6:8]}"
                delivery_dates[cur] = d
            elif seg_name.startswith('E2EDP19') and cur:
                materials[cur] = {'idtnr': data[3:38].strip(), 'ktext': data[38:108].strip()}

        r3 = 4
        for posex, pdata in sorted(positions.items()):
            mat = materials.get(posex, {})
            fill = ALT_FILL if r3 % 2 == 0 else WHT_FILL
            for col, v in enumerate([posex, mat.get('idtnr',''), mat.get('ktext',''),
                                      pdata['menge'], pdata['menee'], pdata['vprei'],
                                      pdata['netwr'], delivery_dates.get(posex,''),
                                      pdata['werks']], 1):
                cell(ws3, r3, col, v, fill)
            r3 += 1

        for col, w in zip('ABCDEFGHI', [8,20,35,12,6,14,14,14,10]):
            ws3.column_dimensions[col].width = w
        ws3.freeze_panes = 'A4'

    wb.save(out_path)
    print(f"Saved: {out_path}")


def build_preview_data(idoc_rows):
    """Return the same data as build_excel but as structured dicts for web preview."""
    rows_out = []
    seg_counter, seg_hlevel, seg_order = {}, {}, {}
    order_idx = 0

    for seg_name, hlevel, data in idoc_rows:
        seg_counter[seg_name] = seg_counter.get(seg_name, 0) + 1
        if seg_name not in seg_hlevel:
            seg_hlevel[seg_name] = hlevel
            seg_order[seg_name] = order_idx
            order_idx += 1
        occ = seg_counter[seg_name]
        fields = extract_fields(seg_name, data)
        first = True
        for fname, flen, fdesc, val in fields:
            meaning = field_meaning(seg_name, fname, val)
            rows_out.append({
                'seg': seg_name if first else '',
                'occ': occ if first else None,
                'lvl': hlevel if first else None,
                'field': fname,
                'len': flen,
                'desc': fdesc,
                'val': val,
                'meaning': meaning,
                'is_first': first,
                'is_empty': not val,
            })
            first = False

    summary = []
    for seg, cnt in sorted(seg_counter.items(), key=lambda x: seg_order.get(x[0], 9999)):
        key = get_key(seg)
        desc = SEG_DESC.get(seg, SEG_DESC.get(key, ''))
        summary.append({
            'seg': seg,
            'cnt': cnt,
            'desc': desc,
            'lvl': seg_hlevel.get(seg, 0),
        })

    positions = []
    if any(row[0].startswith('E2EDP01') for row in idoc_rows):
        pos_map, dates_map, mat_map, cur = {}, {}, {}, None
        for seg_name, hlevel, data in idoc_rows:
            if seg_name.startswith('E2EDP01'):
                posex = data[0:6].strip(); cur = posex
                pos_map[posex] = {
                    'menge': data[10:25].strip(), 'menee': data[25:28].strip(),
                    'vprei': data[35:50].strip(), 'netwr': data[59:74].strip(),
                    'werks': data[193:197].strip(),
                }
            elif seg_name.startswith('E2EDP20') and cur:
                d = data[30:38].strip()
                if len(d) == 8:
                    d = f"{d[:4]}-{d[4:6]}-{d[6:8]}"
                dates_map[cur] = d
            elif seg_name.startswith('E2EDP19') and cur:
                mat_map[cur] = {'idtnr': data[3:38].strip(), 'ktext': data[38:108].strip()}
        for posex, pdata in sorted(pos_map.items()):
            mat = mat_map.get(posex, {})
            positions.append({
                'posex': posex,
                'idtnr': mat.get('idtnr', ''),
                'ktext': mat.get('ktext', ''),
                'menge': pdata['menge'],
                'menee': pdata['menee'],
                'vprei': pdata['vprei'],
                'netwr': pdata['netwr'],
                'date': dates_map.get(posex, ''),
                'werks': pdata['werks'],
            })

    return {'summary': summary, 'rows': rows_out, 'positions': positions}


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("IDoc → Excel dokumentacja")
        print("Użycie:  python idoc_parser.py <plik_idoc.txt> [plik_wyjsciowy.xlsx]")
        print("Przykład: python idoc_parser.py ORDERS_344242684.txt")
        print()
        print("Obsługiwane typy IDoc: ORDERS05, DELFOR02")
        sys.exit(0)
    inp = sys.argv[1]
    if not os.path.exists(inp):
        print(f"Błąd: plik '{inp}' nie istnieje.")
        sys.exit(1)
    out = sys.argv[2] if len(sys.argv) > 2 else inp.replace('.txt', '_dokumentacja.xlsx')
    print(f"Przetwarzam: {inp}")
    build_excel(parse_flat(inp), out)
    print(f"Gotowe:      {out}")
